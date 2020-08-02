import pandas as pd
import os, numpy
from datetime import datetime # anscheinend muss man datetime von datetime importieren, damit man strptime nutzen kann

from openpyxl import load_workbook #for writing in a specific sheet
import itertools

dirfilepath='/media/larfan/ESD-USB/actualwork/input/2020-07-30/AtmoTube/F8073D8A668A-30-Jul-2020-12-58-08.csv'
xlsx='/media/larfan/ESD-USB/actualwork/output/xlsx Arbeitsmappen-2020-07-30/AtmoTube.xlsx'

#open dataframe in pandas
data=pd.read_csv(dirfilepath, sep=',')
names=list(data) #columnnames

#split into intervalls
series=data['Date']
serieslist=series.values.tolist()
for index, item in enumerate(serieslist):
    serieslist[index]=serieslist[index][-8:]

#convert timestamp into datetime format
for t, string in enumerate(serieslist):             #hier benutzen von enumerate, damit man gleichzeitig das listenobjekt, als auch den entsprechenden index durchloopen kann
    serieslist[t]=datetime.strptime(string, '%H:%M:%S')     #damit man die Objekte in der Liste live modifizieren kann
print(serieslist)

count=1

intervalllist=[0]      #-1 muss sein, damit die erste zeile auf der ersten seite gedruckt wird
#weitere Schritte: Array/Dictionary machen aus anfangs und endzelle für ensprechendes Intervall, damit man das dann eintragen kann, Index dann 1-7
for u in range(len(serieslist)):
    print('INDEX: '+ str(u) +'\t'+ str(serieslist[u]))
    try:                                            #try is there for last row when serieslist[u+1] doesnt exist, because u+1 is out of index range
        difference=serieslist[u+1]-serieslist[u]    #das liefert timedelta
        mindifference=difference.total_seconds()/60 #total_seconds() berechnet Sekundenunterschied des timedeltas
        if mindifference>10:
            print('Fahrtzeit')
            intervalllist.append(u+1)               #sehr wichtig ist das #1, sonst funktioniert nichts
            
    except:
        pass

intervalllist.append(len(serieslist))
print(intervalllist)


#zum modifizieren wenn timestamps nicht stimmen:
#intervalllist=[0, 24, 28,42, 78, 92, 106, 120]


book=load_workbook(xlsx)

for i in range(5):      #hier anpassen auf wieviele stationen dann sind(funktioniert nur wenn beispielsweise die beiden letzten stationen nicht angefahren wurden. )
                        #evtl mal lösen mit len(intervallist)!!!



    sheet=book["Stadt%s" % str(i+1)]
    print(sheet.max_column)
    last_column=sheet.max_column


    for column in range(0, len(data.columns)): 

        #TO DO: loop through 2 loops with zip; the existing one and the intervall length e.g intervalllist[i+1]-intervallist[iflu]
        if i<7:                       

            for (row,rowindex) in zip(range(intervalllist[i], intervalllist[i+1]),range(intervalllist[i+1]+1-intervalllist[i])):
                #print('Das ist das Sheet'+str(i)+'und das ist die row:'+str(row)+ ' und das ist der rowindex:'+str(rowindex))
                if rowindex==0:         #das hier ist zweimal geschachtelt und nicht mit if ... and ... ,weil names[] ja in die erste Zeile geschrieben werden muss
                    if column==0:
                        sheet.cell(row=rowindex+2, column=column+2+last_column, value=datetime.today())
                    sheet.cell(row=rowindex+1, column=column+2+last_column, value=names[column])
                sheet.cell(row=rowindex+3, column=column+2+last_column, value=data.iat[row, column])
        

book.save(xlsx)

