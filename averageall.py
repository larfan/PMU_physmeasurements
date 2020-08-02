import openpyxl

specfolder='xlsx Arbeitsmappen-2020-07-30/'
directory='/home/larfan/Documents/PythonProgramming/Excel_csv/allaverage_files/{}'.format(specfolder)
xlsx=directory+'gesmessungenhauptblatt.xlsx'

#get column name in letter format
filesdirectory='=MITTELWERT(\'F:\\allaverage\\{}'.format(specfolder) #MUSS so bleiben, damit man dann die Daten per Stick transferieren kann
def getletcol(x):
    return(openpyxl.utils.get_column_letter(x))


#open writable file
wb=openpyxl.load_workbook(xlsx)
sheet=wb['Tabelle1']

#choose locations
rmlist=[]
locations=['Stadt1','Stadt2','Stadt3','Stadt4','Stadt5','Stadt6','Stadt7']  #derweil noch mit liste, da man so leichter bestimmete locations ausschließen kann
for o in range(len(locations)):
    print(str(o)+': '+locations[o])
while True:
    iput=input('Enter unwanted places (Enter \'exit\' if you are finished!):')
    
    if iput.lower() != 'exit':
        rmlist.append(int(iput))
    else:
        print('we exited')
        print(rmlist)
        break

#choose devices
x=''
rmdevice=[
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],
    [x,x,x,x,x,x,x,x],   
]

list1=[
[2,1,'COM-3200PROII.xlsx',4,8,400],
[3,3,'COM-3200PROII.xlsx',4,2,400],
[6,7,'AtmoTube.xlsx',3,8,30],
[13,11,'PM2.5-Detector.xlsx',4,10,10],
[24,8,'Naneos-partector2.xlsx',4,18,800],
[29,3,'Naneos-partector2.xlsx',4,6,800],
[32,1,'Me-3030B.xlsx',3,0,10],
[33,6,'WheatherFlow-windmeter.xlsx',4,5,6]
]

while True:
    inputloc=input('Enter Index of location with device exceptions! \nIf finished, enter \'exit\':')
    if inputloc.lower() != 'exit':
        for (u,p) in zip(list1,range(len(list1))):
            print(str(p)+' '+u[2])
        while True:
            inputdev=input('Enter device Index, if finished enter \'return\'    :')
            if inputdev.lower() != 'return':
                rmdevice[int(inputloc)][int(inputdev)]='z'
            else:
                print(rmdevice)
                break
    else:
        print('we exited')
        print(rmlist)
        break



for i in rmlist:
    print('das ist i:'+str(i))
    locations[i]='skip'
print('Only the averages of the follwing locations will be written into the Excel file: '+str(locations))
#Liste mit 'Unterlisten', welche für jedes Messgerät die wichtigen Daten enhält:
#[position, length, filename, startrow, distancetolastcolumn,rowlength]



count=0
for dex,loc in enumerate(locations):
    list2=[
    [2,1,'COM-3200PROII.xlsx',4,8,400],
    [3,3,'COM-3200PROII.xlsx',4,2,400],
    [6,7,'AtmoTube.xlsx',3,8,30],
    [13,11,'PM2.5-Detector.xlsx',4,10,10],
    [24,8,'Naneos-partector2.xlsx',4,18,800],
    [29,3,'Naneos-partector2.xlsx',4,6,800],
    [32,1,'Me-3030B.xlsx',3,0,10],
    [33,6,'WheatherFlow-windmeter.xlsx',4,5,6]
    ]
    print(list2)
    print('gehst du hier rein?') #Victor, das ist zum herausfinden, warum er die liste nicht erneuert
    if loc !='skip':
        print('kein skip')
        

        #add skip to device list
        for ind,otem in enumerate(rmdevice[dex]):
                if otem != x:
                    print(ind)
                    list2[ind]='skip' 
                    print('wie oft gehst dur hier rein?')
        print(list2)

        for item in list2:
           
            if item=='skip':
                pass
            else:
                currdata=openpyxl.load_workbook(directory+'{}'.format(item[2]))
                datasheet=currdata[loc]
                lastcol=datasheet.max_column
                print(lastcol)
                currdata.close()


                for length in range(item[1]):
                    
                    #sheet.cell(row=15,column=item[0]+length).value=item[0]

                
                    sheet.cell(row=22,column=item[0]+length+count).value=filesdirectory+'\\[{file}]{stadt}\'!{acolumn}{rowstart}:{acolumn}{endrow})'.format(file=item[2],stadt=loc,acolumn=getletcol(lastcol-item[4]+length),rowstart=item[3],endrow=item[5]) #(getletcol(65),getnumcol(65))
        count+=37
    else:
        print('ungleich skip')
        count+=37
        pass
wb.save(xlsx)
