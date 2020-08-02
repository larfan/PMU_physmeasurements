import pandas as pd
import datetime, os, numpy
from openpyxl import load_workbook #for writing in a specific sheet


#filenames and directories
xlsxfile='/media/larfan/ESD-USB/actualwork/output/xlsx Arbeitsmappen-2020-07-30/COM-3200PROII.xlsx'
datadirectory='/media/larfan/ESD-USB/actualwork/input/2020-07-30/Com-3200/'

names=['Messung','Ionenart','/CC','C°','Humi%'] #columnnames


#grab last 4 characters of the file name:
def last_4chars(x):
    return(x[5:11])

#get rid of " marks
def remover(currentfile,): 
    f1=open(currentfile, 'r')
    f2=open('tempfile.csv', 'w')
    for row in f1:
            row = row.strip() # remove "\n"
            row = row[1:-1] # remove " on both ends
            f2.write(row + '\n')
    f2.close()
    f1.close()

#get first line(just needs to be called evertime, as to not stay the same value)
def firstline():
    f2=open('tempfile.csv', 'r')
    allines=f2.readlines()
    global messung
    messung=allines[0]
    messung.strip()
    f2.close()
    
#get dirfiles
dirfiles=[]
for f in os.listdir(datadirectory): #for loop removes hidden files in directory
        if not f.startswith('.'):
            dirfiles.append(f)
dirfiles=sorted(dirfiles, key=last_4chars)
print(dirfiles)

        


book=load_workbook(xlsxfile)
count=1

for i in range(14):
    
    #get rid of " marks
    thisfile=datadirectory + str(dirfiles[i])
    remover(thisfile)

    #Bekommen von erster Zeile(unsauber weil mans nocheinmal aufmachen muss;funkioniert aber sonst nicht)
    firstline()

    #define right sheet
    data=pd.read_csv("tempfile.csv",skiprows=1,sep='\t')
    if i % 2 == 0:
        sheet=book["Stadt" + str(count)]
        print("Stadt" + str(count))
    else:
        pass
        count=count+1

    last_column=sheet.max_column


    for column in range(0, len(data.columns)):      
        for row in range(0, len(data.index)):
            if row==0:
                if column==0:
                    sheet.cell(row=row+2, column=column+2+last_column, value=datetime.date.today())  
                    sheet.cell(row=row+3, column=column+2+last_column, value=messung)               
                sheet.cell(row=row+1, column=column+2+last_column, value=names[column])
            else:
                sheet.cell(row=row+3, column=column+2+last_column, value=data.iat[row-1, column])                  #df.iat() kann zellen aus dataframe mit integer angeben
print('(Hopefully) All files were copied to their respective locations')
book.save(xlsxfile)


