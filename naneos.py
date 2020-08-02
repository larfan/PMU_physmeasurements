import pandas as pd
import os, numpy
from datetime import datetime
from openpyxl import load_workbook #for writing in a specific sheet

xlsx='/media/larfan/ESD-USB/actualwork/output/xlsx Arbeitsmappen-2020-07-30/Naneos-partector2.xlsx'
directory='/media/larfan/ESD-USB/actualwork/input/2020-07-30/naneos-partector-2/'

#grab last 4 characters of the file name:
def last_4chars(x):
    return(x[5:11])

#get dirfiles
dirfiles=[]
for f in os.listdir(directory): #for loop removes hidden files in directory
        if not f.startswith('.'):
            dirfiles.append(f)
dirfiles=sorted(dirfiles, key=last_4chars)
print(dirfiles)


#rows=dataframe_to_rows(data)

book=load_workbook(xlsx)

for i in range(7):

    #open dataframe in pandas and get the 8th line of document
    data=pd.read_csv(directory+dirfiles[i], sep='\t',skiprows=18)
    lines=open(directory+ str(dirfiles[i]))
       
    
    allines=lines.readlines()
    linedate=allines[7]
    print(linedate)
    names=list(data) #columnnames

    sheet=book["Stadt%s" % str(i+1)]
    print(sheet.max_column)
    last_column=sheet.max_column


    for column in range(0, len(data.columns)):      
        for row in range(0, len(data.index)):
            if row==0:
                if column==0:
                    sheet.cell(row=row+2, column=column+2+last_column, value=datetime.today())
                    sheet.cell(row=row+3, column=column+2+last_column, value=linedate)
                print(list(data))
                sheet.cell(row=row+1, column=column+2+last_column, value=names[column])
                sheet.cell(row=row+4, column=column+2+last_column, value=data.iat[row, column]) #df.iat() kann zellen aus dataframe mit integer angeben
            else:
                #print('1')
                sheet.cell(row=row+4, column=column+2+last_column, value=data.iat[row, column])
book.save(xlsx)

